import os
from openpyxl import Workbook
import pdb
 

from openpyxl import load_workbook
 
# 优化后类型
workbook1 = load_workbook(filename='attribute4.xlsx')
sheet1 = workbook1.active

#原始类型
workbook2 = load_workbook(filename='state_var_result.xlsx')
sheet2 = workbook2.active

#match表
workbook3 = load_workbook(filename='state_match.xlsx')
sheet3 = workbook3.active

optim_col=[cell.value for cell in sheet1['A']]
source_col=[cell.value for cell in sheet2['A']]
match_col=[cell.value for cell in sheet3['A']]

error=0
correct=0
n=0
a=0
b=0
c=0
with open('./attribute_error.txt','w') as f:
    for i in range(1,len(optim_col)):
        contract=optim_col[i]
        #if contract[:-4]!='0x20dfe8004103cfd78de46544c938ef99c2737dee':
        #    continue
        if contract not in source_col or contract not in match_col:
            continue
        source_index=source_col.index(contract)
        match_index=match_col.index(contract)
        
        optim_row=sheet1[i+1]
        source_row=sheet2[source_index+1]
        match_row=sheet3[match_index+1]
        
        optime_dict={}
        source_dict={}
        match_dict={}
        
        for j in range(1,len(optim_row),2):
            if optim_row[j].value==None or optim_row[j+1].value==None:
                break
            optime_dict.update({optim_row[j].value.strip():optim_row[j+1].value.strip()})     
        
        for j in range(3,len(source_row)):
            if source_row[j].value==None :
                break
            cell1=source_row[j].value
            if ':' not in cell1:
                continue
            
            cell1=cell1.split(':')
        
            var=cell1[0].strip()
            attribute=cell1[1].strip()
            
            if 'bots' in var or 'preTrader' in var or 'isExcludedFromFee' in var:
                attribute='flag'
            
                
            source_dict.update({var:attribute})   
            
        for j in range(1,len(match_row),2):
            if match_row[j].value==None or  match_row[j+1].value==None:
                break
            match_dict.update({match_row[j].value.strip():match_row[j+1].value.strip()})      
            
        for source_variable in match_dict:
            optim_variable=match_dict[source_variable]
            #print(optim_variable)
            
            
            if optim_variable in optime_dict and source_variable in source_dict:
                #print(contract+' : 源码 : '+source_variable+' : '+source_dict[source_variable]+'.   反编译 : '+ optim_variable+' : '+ optime_dict[optim_variable]+'\n')
                if source_dict[source_variable]=='others' or source_variable=='_implementation':
                    a+=1
                    continue
                if optime_dict[optim_variable]==source_dict[source_variable]:
                    if source_dict[source_variable]=='others':
                        c+=1
        
                    correct+=1
                else:
                    if 'uniswapV2'in  optim_variable and optime_dict[optim_variable]=='router':
                        correct+=1
                    #elif optime_dict[optim_variable]=='flag' and source_dict[source_variable]=='limit':
                    #    correct+=1
                    #elif optime_dict[optim_variable]=='limit' and source_dict[source_variable]=='flag':
                    #    correct+=1
                    else:
                        if source_dict[source_variable]=='others':
                            b+=1
                        if optime_dict[optim_variable]=='others':
                            n+=1
                        
                        f.write(contract+' : 源码 : '+source_variable+' : '+source_dict[source_variable]+'.   反编译 : '+ optim_variable+' : '+ optime_dict[optim_variable]+'\n')
                        error+=1
        
            
print('correct:{}   error:{}   rate:{}'.format(correct,error,correct/(correct+error)))