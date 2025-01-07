import os
from openpyxl import Workbook
import pdb
 

from openpyxl import load_workbook
 
# 优化后类型
workbook1 = load_workbook(filename='optim_type_match4.xlsx')
sheet1 = workbook1.active

#原始类型
workbook2 = load_workbook(filename='source.xlsx')
sheet2 = workbook2.active

#match表
workbook3 = load_workbook(filename='state_match.xlsx')
sheet3 = workbook3.active

optim_col=[cell.value for cell in sheet1['A']]
source_col=[cell.value for cell in sheet2['A']]
match_col=[cell.value for cell in sheet3['A']]

error=0
correct=0
a=0
for i in range(1,len(optim_col)):
    contract=optim_col[i]
    if contract not in source_col or contract not in match_col:
        continue
    source_index=source_col.index(contract)
    match_index=match_col.index(contract)
    
    optim_row=sheet1[i+1]
    source_row=sheet2[source_index+1]
    match_row=sheet3[match_index+1]
    
    optime_dict={}
    source_dict={}
    match_dict={}
    
    for j in range(1,len(optim_row),2):
        if optim_row[j].value==None  or optim_row[j+1].value==None:
            break
        optime_dict.update({optim_row[j].value.strip():optim_row[j+1].value.strip()})     
    
    for j in range(1,len(source_row),2):
        if source_row[j].value==None or source_row[j+1].value==None:
            break
        source_dict.update({source_row[j].value.strip():source_row[j+1].value.strip()})   
        
    for j in range(1,len(match_row),2):
        if match_row[j].value==None or  match_row[j+1].value==None:
            break
        match_dict.update({match_row[j].value.strip():match_row[j+1].value.strip()})      
        
    for source_variable in match_dict:
        #if source_variable=='_implementation' or source_variable=='_IMPLEMENTATION_SLOT':
        #    continue
        optim_variable=match_dict[source_variable]
        if optim_variable in optime_dict and source_variable in source_dict:
            
            if optime_dict[optim_variable]==source_dict[source_variable]:
                correct+=1
            else:
                if optime_dict[optim_variable].startswith('mapping') and source_dict[source_variable].startswith('mapping'):
                    correct+=1
                #elif 'byte' in optime_dict[optim_variable] and 'string' in source_dict[source_variable]:
                #    correct+=1
                #elif 'byte' in  source_dict[source_variable] and 'string' in optime_dict[optim_variable]:
                #    correct+=1
                #elif source_dict[source_variable].startswith('uint') and optime_dict[optim_variable].startswith('uint'):
                #    correct+=1
                elif source_dict[source_variable].startswith('string') and optime_dict[optim_variable].startswith('byte'):
                    correct+=1
                #elif source_dict[source_variable].startswith('byte') and optime_dict[optim_variable].startswith('string'):
                #    correct+=1
                else:
                    with open('result.txt','a+') as f:
                        f.writelines(contract+' : 源码 : '+source_variable+' : '+source_dict[source_variable]+'.   反编译 : '+ optim_variable+' : '+ optime_dict[optim_variable]+'\n')
                    error+=1
        else:
            
            #print(contract+' : 源码 : '+source_variable+'.   反编译 : '+ optim_variable+' : ')
            a+=1
            
print('correct:{}   error:{}   rate:{}'.format(correct,error,correct/(correct+error)))
print(a)