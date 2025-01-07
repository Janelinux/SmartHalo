import os
from openpyxl import Workbook
import pdb
# 创建一个新的工作簿
wb = Workbook()
 
# 选择默认的工作表
ws = wb.active
 

from openpyxl import load_workbook
 
# 加载现有的Excel文件
workbook = load_workbook(filename='storage_decom.xlsx')
sheet = workbook.active
 
# 获取第一列的标题（例如：'A'）
column_values = [cell.value for cell in sheet['A']]
#print(column_values)
ws.append(["合约","反编译码中状态变量","反编译码中状态变量类型","以此类推...."])
for adir in os.listdir('./compare_function_boroder'):
    print(adir)
    #if not adir=='0x84003ede8965ed05bec1bda85f12658b9803a5d5':
    #    continue
    adir_path=os.path.join('./compare_function_boroder',adir)
    for file in os.listdir(adir_path):
        file_path=os.path.join(adir_path,file)
        if file.startswith('optim_'):
            name_type={}
            with open(file_path,'r') as f:
                lines=f.readlines()
                    
            for line in lines:
                annotes=line.strip().split('//')[1:]
                line=line.strip().split('//')[0]
                if ("private " in line or "public " in line or "constant " in line or "internal " in line or "mapping" in line or line.strip().startswith('uint') or line.strip().startswith('address')) and "function " not in line and "*" not in line and '{' not in line :
                    #pdb.set_trace()
                    line=line.split(";")
                    for str in line:
                        if  "private" in str or "public" in str or "constant" in str or "internal" in str  or "mapping" in line or str.strip().startswith('uint') or str.strip().startswith('address'):
                            if "mapping" in str:
                                str=str.split()
                                name=''
                                for annote in annotes:
                                    if 'The name of the corresponding state variable before the modification' in annote or 'The corresponding state variable' in annote:
                                        name=annote.strip().split(':')[1]
                                        name=name.strip()
                                        break
                                    elif 'STORAGE' in annote:
                                        index=column_values.index(adir+'.sol')+1
                                        row=sheet[index]
                                        storage=annote.strip()
                                        
                                        for i in range(1,len(row),2):
                                            if row[i].value==None:
                                                break
                                            if row[i+1].value.strip()==storage:
                                                name=row[i].value.strip()
                                                break      
                                                            
                                if name=='':
                                    continue
                                name_dict={name:' '.join(str[:-2]).strip()}
                                name_type.update(name_dict)
                            else:
                                str=str.split("=")
                                for str2 in str:
                                    if  "private" in str2 or "public" in str2 or "constant" in str2 or "internal" in str2 or str2.strip().startswith('uint') or str2.strip().startswith('address'):
                                        str2=str2.split()
                                        for annote in annotes:
                                            if 'The name of the corresponding state variable before the modification' in annote or 'The corresponding state variable' in annote:
                                                if ':' in annote:
                                                    name=annote.strip().split(':')[1]
                                                    name=name.strip()
                                                break
                                            elif 'STORAGE' in annote:
                                                index=column_values.index(adir+'.sol')+1
                                                row=sheet[index]
                                                storage=annote.strip()
                                        
                                                for i in range(1,len(row),2):
                                                    #print(row[i].value)
                                                    if row[i].value==None:
                                                        break
                                                    if row[i+1].value.strip()==storage:
                                                        name=row[i].value.strip()
                                                        break   
                                                
                                        if name=='':
                                            continue
                                        name_dict={name:str2[0].strip()}
                                        name_type.update(name_dict)
            all_var=[]
            all_var.append(adir+'.sol')
            for key in name_type:
                all_var.append(key)
                all_var.append(name_type[key])
            ws.append(all_var)
wb.save('optim_type_match.xlsx')
                